#!/usr/bin/env python3
"""Notifie sur Telegram les nouvelles lignes des onglets « EOW Console »,
« EOD Console », « Idees Console » et « EOD Setter Console » du classeur closing
(remplis par la console via le pont).

Usage: python3 eow_notify.py investisseurs30.xlsx ../.eow-count [../.eod-count] [../.idea-count] [../.eodsetter-count]
Env : TELEGRAM_TOKEN, TELEGRAM_CHAT (sinon ne fait rien).
Chaque fichier d'état contient le nombre de lignes déjà notifiées (commité par
le workflow pour persister entre les runs).
"""
import json
import os
import sys
import urllib.parse
import urllib.request

import openpyxl


def send(token, chat, text):
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=urllib.parse.urlencode({"chat_id": chat, "text": text}).encode(),
        method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        out = json.load(resp)
    if not out.get("ok"):
        raise RuntimeError(f"Telegram: {out}")


def fmt_eow(vals):
    date, closer, energie, leads, good, hard, obj, need, free = (vals + [""] * 9)[:9]
    return (f"EOW reçu de {closer or '?'}\n\n"
            f"Énergie : {energie or '?'}/10 · Ressenti leads : {leads or '?'}/10\n\n"
            f"Ce qui a marché : {good or '·'}\n"
            f"Ce qui a bloqué : {hard or '·'}\n"
            f"Objection : {obj or '·'}\n"
            f"Besoin équipe : {need or '·'}\n"
            f"Autre : {free or '·'}")


def fmt_eod(vals):
    date, closer, energie, mood, flag = (vals + [""] * 5)[:5]
    return (f"🌙 EOD reçu de {closer or '?'}\n\n"
            f"Énergie : {energie or '?'}/10\n"
            f"Journée : {mood or '·'}\n"
            f"À signaler : {flag or '·'}")


def fmt_eod_setter(vals):
    ts, jour, setter, energie, acc, rel, rep, quali, prop, book, obj, ret, aud, con = (vals + [""] * 14)[:14]
    j = jour[8:10] + "/" + jour[5:7] if len(jour) >= 10 and jour[4] == "-" else (jour or "?")
    return (f"📱 EOD setting reçu de {setter or '?'} ({j})\n\n"
            f"Énergie : {energie or '?'}/10\n"
            f"Audits reçus : {aud or '0'} · Leads audit contactés : {con or '0'}\n"
            f"Conv. qualifiées : {quali or '0'} · RDV proposés : {prop or '0'} · RDV bookés : {book or '0'}\n"
            f"Accroches : {acc or '0'} · Relances : {rel or '0'}\n\n"
            f"Objections / blocages : {obj or '·'}\n"
            f"Retours : {ret or '·'}")


def fmt_idea(vals):
    date, closer, idea = (vals + [""] * 3)[:3]
    return f"💡 Idée de {closer or '?'}\n\n{idea or '·'}"


def notify_tab(wb, token, chat, tab, state_path, fmt, label):
    if tab not in wb.sheetnames:
        print(f"onglet {tab} absent (rien d'envoyé pour l'instant)", file=sys.stderr)
        return
    rows = [r for r in wb[tab].iter_rows(values_only=True) if any(r)]
    n = len(rows) - 1  # moins l'en-tête
    try:
        seen = int(open(state_path).read().strip())
    except Exception:
        seen = 0
    if n <= seen:
        print(f"{label} : rien de nouveau ({n} au total)", file=sys.stderr)
        return
    for r in rows[1 + seen:]:
        vals = [str(v).strip() if v is not None else "" for v in r]
        # les floats du Sheet (7.0) s'affichent en entier
        vals = [v[:-2] if v.endswith(".0") else v for v in vals]
        send(token, chat, fmt(vals))
        print(f"{label} notifié : {vals[1] if len(vals) > 1 else '?'}", file=sys.stderr)
    with open(state_path, "w") as f:
        f.write(str(n))


def main(xlsx_path, eow_state, eod_state=None, idea_state=None, setter_state=None):
    token = os.environ.get("TELEGRAM_TOKEN")
    chat = os.environ.get("TELEGRAM_CHAT")
    if not (token and chat):
        print("pas de config Telegram, EOW/EOD non notifiés", file=sys.stderr)
        return
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    notify_tab(wb, token, chat, "EOW Console", eow_state, fmt_eow, "EOW")
    if eod_state:
        notify_tab(wb, token, chat, "EOD Console", eod_state, fmt_eod, "EOD")
    if idea_state:
        notify_tab(wb, token, chat, "Idees Console", idea_state, fmt_idea, "Idée")
    if setter_state:
        notify_tab(wb, token, chat, "EOD Setter Console", setter_state, fmt_eod_setter, "EOD setter")


if __name__ == "__main__":
    main(*sys.argv[1:6])
