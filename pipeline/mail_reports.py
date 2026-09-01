#!/usr/bin/env python3
"""Mails automatiques aux closers, envoyés par le pont Apps Script (what=send_mail,
depuis la boîte Gmail d'Alex).

Usage:
  python3 mail_reports.py eow  [--dry-run]   # lundi matin : relance les closers sans EOW
  python3 mail_reports.py fill [--dry-run]   # chaque soir : recap des calls non remplis

Entrées :
  - data.json (parse_xlsx) dans le dossier courant
  - investisseurs30.xlsx (onglet « EOW Console ») pour le job eow
Config :
  - env BRIDGE_URL / BRIDGE_KEY (sinon bridge.json local)
  - env CLOSER_EMAILS = JSON {"Louis Granier": "..."} (sinon closers.json local)
  - env MAIL_ALEX (copie recap à Alex, défaut alexyoucompte99@gmail.com)
Un mail par closer concerné + un recap à Alex. Aucun mail si rien à signaler.
"""
import datetime as dt
import json
import os
import sys
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
CONSOLE_URL = "https://alexyoucompte99-lang.github.io/AlexI3/closing/"
TOP3 = ["Louis Granier", "Romain Gourand", "Adrien Soret"]


def bridge_cfg():
    url, key = os.environ.get("BRIDGE_URL"), os.environ.get("BRIDGE_KEY")
    if not (url and key):
        b = json.load(open(os.path.join(HERE, "bridge.json")))
        url, key = b["url"], b["key"]
    return url, key


def closer_emails():
    raw = os.environ.get("CLOSER_EMAILS")
    if not raw:
        raw = open(os.path.join(HERE, "closers.json")).read()
    return json.loads(raw)


def send_mail(to, subject, body, dry):
    # les logs GitHub Actions du repo public ne doivent contenir ni adresses ni
    # noms de prospects : le contenu complet ne s'affiche qu'en dry-run LOCAL
    if dry:
        if os.environ.get("GITHUB_ACTIONS"):
            print(f"--- [dry] mail prêt ({subject.split('·')[0].strip()}, {len(body)} car.)")
        else:
            print(f"--- mail -> {to} · {subject}\n{body}\n")
        return
    url, key = bridge_cfg()
    payload = json.dumps({"key": key, "what": "send_mail", "to": to,
                          "subject": subject, "body": body}).encode()
    # POST vers le pont : --data sans method forcé (le 302 Apps Script rejoue
    # un GET sinon) ; urllib suit le 302 en GET tout seul, c'est le comportement voulu
    req = urllib.request.Request(url, data=payload)
    with urllib.request.urlopen(req, timeout=60) as resp:
        out = resp.read().decode()
    if '"ok":true' not in out.replace(" ", ""):
        raise RuntimeError(f"pont send_mail: {out[:300]}")
    print(f"✓ mail envoyé ({subject.split('·')[0].strip()})")


def load_calls():
    d = json.load(open(os.path.join(HERE, "data.json")))
    return d["calls"]


def paris_now():
    # les runners GitHub sont en UTC ; on raisonne en heure de Paris
    import zoneinfo
    return dt.datetime.now(zoneinfo.ZoneInfo("Europe/Paris"))


def fr_date(iso):
    y, m, d = iso.split("-")
    return f"{d}/{m}"


# ---------------------------------------------------------------- job EOW
def job_eow(dry):
    now = paris_now()
    today = now.date()
    last_monday = today - dt.timedelta(days=today.weekday() + 7)
    last_sunday = last_monday + dt.timedelta(days=6)

    calls = load_calls()
    actifs = set()
    for c in calls:
        d = c.get("date")
        if d and last_monday.isoformat() <= d <= last_sunday.isoformat():
            name = " ".join((c.get("closer") or "").split())
            if name in TOP3:
                actifs.add(name)

    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(HERE, "investisseurs30.xlsx"),
                                read_only=True, data_only=True)
    done = set()
    if "EOW Console" in wb.sheetnames:
        for r in wb["EOW Console"].iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            when, closer = r[0], " ".join(str(r[1] or "").split())
            if isinstance(when, dt.datetime):
                when = when.date()
            elif isinstance(when, str):
                try:
                    when = dt.date.fromisoformat(when[:10])
                except ValueError:
                    continue
            # EOW compté pour la semaine passée s'il date de cette semaine-là
            # ou du lundi matin courant
            if last_monday <= when <= today and closer in TOP3:
                done.add(closer)

    manquants = [n for n in TOP3 if n in actifs and n not in done]
    emails = closer_emails()
    sem = f"semaine du {last_monday.strftime('%d/%m')} au {last_sunday.strftime('%d/%m')}"
    for name in manquants:
        to = emails.get(name)
        if not to:
            print(f"! pas d'email pour {name}, mail sauté")
            continue
        p = name.split()[0]
        body = (f"Salut {p},\n\n"
                f"Petit rappel du lundi matin : ton EOW de la {sem} n'est pas encore rempli.\n\n"
                f"Ça prend 2 minutes dans l'onglet EOW de la console :\n{CONSOLE_URL}\n\n"
                f"Merci !\nAlex")
        send_mail(to, f"Ton EOW de la semaine ✍️ ({sem})", body, dry)

    alex = os.environ.get("MAIL_ALEX", "alexyoucompte99@gmail.com")
    if manquants:
        recap = (f"Rappel EOW du lundi · {sem}\n\n"
                 f"Closers actifs : {', '.join(sorted(actifs)) or 'aucun'}\n"
                 f"EOW reçus : {', '.join(sorted(done & actifs)) or 'aucun'}\n"
                 f"Relancés par mail : {', '.join(manquants)}")
        send_mail(alex, f"EOW manquants relancés : {', '.join(n.split()[0] for n in manquants)}", recap, dry)
    else:
        print(f"EOW : rien à relancer (actifs: {sorted(actifs)}, reçus: {sorted(done)})")


# ---------------------------------------------------------------- job fill
def fill_todo(calls, now):
    """Même règle que l'onglet « À remplir » de la console : 30 derniers jours,
    call passé (heure comparée pour aujourd'hui), show-up vide ou présent sans
    résultat."""
    today = now.date().isoformat()
    hhmm = now.strftime("%H:%M")
    frm = (now.date() - dt.timedelta(days=30)).isoformat()
    out = []
    for c in calls:
        d = c.get("date")
        if not d or d < frm or d > today:
            continue
        if d == today and not (c.get("hour") and c["hour"] <= hhmm):
            continue
        s, v = (c.get("show_up") or "").upper(), (c.get("vente") or "").strip()
        if not s or (s == "OUI" and not v):
            name = " ".join((c.get("closer") or "").split())
            if name in TOP3:
                out.append((name, d, c.get("hour") or "", (c.get("prospect") or "").strip() or "(sans nom)"))
    return out


def job_fill(dry):
    now = paris_now()
    todo = fill_todo(load_calls(), now)
    emails = closer_emails()
    today = now.date().isoformat()
    alex = os.environ.get("MAIL_ALEX", "alexyoucompte99@gmail.com")
    recap_lines = []
    for name in TOP3:
        mine = sorted([t for t in todo if t[0] == name], key=lambda t: (t[1], t[2]), reverse=True)
        if not mine:
            continue
        recap_lines.append(f"{name} : {len(mine)} call(s) non rempli(s)")
        to = emails.get(name)
        if not to:
            print(f"! pas d'email pour {name}, mail sauté")
            continue
        p = name.split()[0]
        auj = [t for t in mine if t[1] == today]
        anciens = [t for t in mine if t[1] != today]
        lines = [f"Salut {p},", "",
                 f"Il te reste {len(mine)} call(s) à remplir dans le tracking :", ""]
        if auj:
            lines.append("Aujourd'hui :")
            lines += [f"  · {t[2] or '?'} {t[3]}" for t in auj]
        if anciens:
            if auj:
                lines.append("Plus anciens :")
            shown = anciens[:12]
            lines += [f"  · {fr_date(t[1])} {t[3]}" for t in shown]
            if len(anciens) > 12:
                lines.append(f"  · et {len(anciens) - 12} autre(s) plus ancien(s)")
        lines += ["", "Tout se remplit en 2 minutes depuis l'onglet « À remplir » de la console :",
                  CONSOLE_URL, "", "Merci !", "Alex"]
        body = "\n".join(l for l in lines if l is not None)
        send_mail(to, f"Tes calls à remplir · {now.strftime('%d/%m')}", body, dry)
    if recap_lines:
        send_mail(alex, f"Calls non remplis relancés · {now.strftime('%d/%m')}",
                  "Recap du soir envoyé aux closers :\n\n" + "\n".join(recap_lines), dry)
    else:
        print("fill : tout est rempli, aucun mail")


def main():
    dry = "--dry-run" in sys.argv
    job = next((a for a in sys.argv[1:] if not a.startswith("-")), "")
    if job == "eow":
        job_eow(dry)
    elif job == "fill":
        job_fill(dry)
    else:
        sys.exit("usage: mail_reports.py eow|fill [--dry-run]")


if __name__ == "__main__":
    main()
