#!/usr/bin/env python3
"""Alerte Telegram du soir : l'EOD setting du jour (Constant) n'est pas dans le Sheet.

Usage: python3 eod_setter_check.py investisseurs30.xlsx [--dry-run]
Env : TELEGRAM_TOKEN / TELEGRAM_CHAT (sinon telegram.json local {"token","chat_id"}),
      EOD_SETTERS = prénoms attendus séparés par des virgules (défaut « Constant »).
Lit l'onglet « EOD Setter Console » (rempli par l'onglet « EOD Constant » de la
console via le pont v12). Un message par setter attendu SANS ligne datée
d'aujourd'hui (heure de Paris). Si l'EOD est là : rien (la notif « reçu » est
déjà partie via eow_notify.py au refresh). Code retour 0 dans tous les cas.
"""
import datetime as dt
import json
import os
import sys
import urllib.parse
import urllib.request
from zoneinfo import ZoneInfo

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PARIS = ZoneInfo("Europe/Paris")
CONSOLE_URL = "https://alexyoucompte99-lang.github.io/AlexI3/closing/"


def telegram_cfg():
    token, chat = os.environ.get("TELEGRAM_TOKEN"), os.environ.get("TELEGRAM_CHAT")
    if not (token and chat):
        try:
            t = json.load(open(os.path.join(HERE, "telegram.json")))
            token, chat = t["token"], t["chat_id"]
        except Exception:
            return None, None
    return token, chat


def send(token, chat, text):
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=urllib.parse.urlencode({"chat_id": chat, "text": text}).encode(), method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        out = json.load(resp)
    if not out.get("ok"):
        raise RuntimeError(f"Telegram: {out}")


def day_of(v):
    """Colonne « Jour » (texte AAAA-MM-JJ) ou « Date » (datetime) -> date."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v or "").strip()
    try:
        return dt.date.fromisoformat(s[:10])
    except ValueError:
        return None


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "investisseurs30.xlsx")
    dry = "--dry-run" in sys.argv
    setters = [s.strip() for s in os.environ.get("EOD_SETTERS", "Constant").split(",") if s.strip()]
    today = dt.datetime.now(PARIS).date()

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    done, last = {}, {}
    if "EOD Setter Console" in wb.sheetnames:
        rows = [r for r in wb["EOD Setter Console"].iter_rows(values_only=True) if any(r)]
        for r in rows[1:]:
            r = list(r) + [None] * 4
            setter = str(r[2] or "").strip()
            if not setter or setter.lower().startswith("test"):
                continue
            d = day_of(r[1]) or day_of(r[0])
            if not d:
                continue
            key = setter.split(" ")[0].lower()
            last[key] = max(last.get(key, d), d)
            if d == today:
                done[key] = r[0].strftime("%H:%M") if isinstance(r[0], dt.datetime) else ""
    else:
        print("onglet EOD Setter Console absent", file=sys.stderr)

    token, chat = telegram_cfg()
    for s in setters:
        key = s.lower()
        if key in done:
            print(f"{s} : EOD du {today:%d/%m} reçu ({done[key] or 'heure inconnue'}), pas d'alerte")
            continue
        prev = last.get(key)
        txt = (f"⚠️ EOD setting PAS FAIT aujourd'hui ({today:%d/%m}) : {s}\n\n"
               + (f"Dernier EOD reçu : {prev:%d/%m/%Y}" + (f" (il y a {(today - prev).days} j)" if (today - prev).days else "")
                  if prev else "Aucun EOD reçu pour l'instant.")
               + f"\n\nÀ remplir dans l'onglet « EOD Constant » de la console :\n{CONSOLE_URL}")
        if dry or not token:
            print(("[dry] " if dry else "[pas de config Telegram] ") + txt.replace("\n", " | "))
        else:
            send(token, chat, txt)
            print(f"{s} : alerte Telegram envoyée")
    return 0


if __name__ == "__main__":
    sys.exit(main())
